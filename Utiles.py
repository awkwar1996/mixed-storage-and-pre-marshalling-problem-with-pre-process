#!/usr/bin/env/python

'''
Loading data from xlsx files
'''
import math
from itertools import product
import numpy as np
import openpyxl
import torch

def load_data(url):
    '''Load basic information'''
    basic_datasheet = openpyxl.load_workbook(url)['Basic']
    N = int(basic_datasheet.cell(1, 2).value)
    S = int(basic_datasheet.cell(2, 2).value)
    H = int(basic_datasheet.cell(3, 2).value)
    P = int(basic_datasheet.cell(4, 2).value)
    N_i = int(basic_datasheet.cell(5, 2).value)
    M = int(basic_datasheet.cell(6, 2).value)
    R = int(basic_datasheet.cell(7, 2).value)
    W = int(basic_datasheet.cell(8, 2).value)
    eta = float(basic_datasheet.cell(9, 2).value)
    time_for_row_empty = float(basic_datasheet.cell(10, 2).value)
    time_for_column_empty = float(basic_datasheet.cell(11, 2).value)
    time_for_row_loaded = float(basic_datasheet.cell(12, 2).value)
    time_for_column_loaded = float(basic_datasheet.cell(13, 2).value)

    '''Load inbound plate information'''
    inbound_plate_group, preprocess_time = [], []
    inbound_datasheet = openpyxl.load_workbook(url)['Inbound']
    for i in range(M):
        inbound_plate_group.append(int(inbound_datasheet.cell(i+2, 2).value))
        preprocess_time.append(float(inbound_datasheet.cell(i+2, 3).value))

    '''Load storage information'''
    storage_configuration = [[] for s in range(S)]
    configuration_datasheet = openpyxl.load_workbook(url)['Configuration']
    for s in range(S):
        for h in range(H):
            if configuration_datasheet.cell(s+1, h+1).value is None or configuration_datasheet.cell(s+1, h+1).value == "":
                break
            storage_configuration[s].append(int(configuration_datasheet.cell(s+1, h+1).value))

    '''Calculate the time for one operate'''
    operate_time = np.zeros((S+1, S+1, S+1), dtype=float)
    min_operate_time = math.inf
    for s1, s2, s3 in product(range(S+1), range(S+1), range(S+1)):
        # row, column for s1, s2, s3
        s1_row = 1 if s1 == 0 else math.ceil(s1/W)
        s1_col = 0 if s1 == 0 else s1 - (s1_row - 1) * W
        s2_row = 1 if s2 == 0 else math.ceil(s2 / W)
        s2_col = 0 if s2 == 0 else s2 - (s2_row - 1) * W
        s3_row = 1 if s3 == 0 else math.ceil(s3 / W)
        s3_col = 0 if s3 == 0 else s3 - (s3_row - 1) * W
        #calculate time
        operate_time[s1, s2, s3] = 2 * eta + \
                                   abs(s1_row - s2_row) * (time_for_row_loaded + time_for_row_empty) + \
                                   abs(s1_col - s2_col) * (time_for_column_loaded + time_for_column_empty) + \
                                   abs(s2_row - s3_row) * (time_for_row_loaded + time_for_row_empty) + \
                                   abs(s2_col - s3_col) * (time_for_column_loaded + time_for_column_empty)
        if s2 != s3:
            min_operate_time = min(min_operate_time, operate_time[s1, s2, s3])
    return S, H, M, N_i, P, inbound_plate_group, preprocess_time, storage_configuration, operate_time, min_operate_time
def parameter_setting():
    settings = {
    # parameters for aco
    'Ant_number': 30,
    'aco_Iteration': 600,
    'weight_decay': 0.95,
    #parameters for instances generate
    'S': 20,
    'H': 20,
    'occ': 0.8,
    'M_list': [10, 15, 20, 25, 30],
    'mean_list': [20, 25, 30, 35, 40],
    'std_list': [2],
    'probability': [0.04, 0.06, 0.11, 0.14, 0.18, 0.22, 0.25],
    'P': 7,
    # parameters for GA
    'population': 100,
    'GA_iteration': 600,
    'mutation_rate':0.05,
    'crossover_rate': 0.95
    }
    return settings

def save_result(result_list, workbook_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in range(len(result_list)):
        sheet.cell(row=i + 1, column=1).value = i
        sheet.cell(row=i + 1, column=2).value = result_list[i]
    workbook.save(f'{workbook_name}.xlsx')
def save_result1(result_list, workbook_name, time):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = 'time'
    sheet.cell(row=1, column=2).value = time
    for i in range(len(result_list)):
        sheet.cell(row=i + 2, column=1).value = i
        sheet.cell(row=i + 2, column=2).value = result_list[i]
    workbook.save(f'{workbook_name}.xlsx')

def save_result2(result_list, relTime_list, bpNum_list, workbook_name, time):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1).value = 'time'
    sheet.cell(row=1, column=2).value = time
    sheet.cell(row=2, column=1).value = 'iteration'
    sheet.cell(row=2, column=2).value = 'result'
    sheet.cell(row=2, column=3).value = 'in-yard relocation time'
    sheet.cell(row=2, column=4).value = 'bp number'
    for i in range(len(result_list)):
        sheet.cell(row=i + 3, column=1).value = i
        sheet.cell(row=i + 3, column=2).value = result_list[i]
        sheet.cell(row=i + 3, column=3).value = relTime_list[i]
        sheet.cell(row=i + 3, column=4).value = bpNum_list[i]
    workbook.save(f'{workbook_name}.xlsx')

if __name__ == '__main__':
    pass

