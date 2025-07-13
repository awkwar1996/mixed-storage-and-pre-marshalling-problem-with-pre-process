from mpl_toolkits.mplot3d import Axes3D

from LA_N import extended_LA_for_heu
import random
import pandas as pd
import openpyxl
import os
from sklearn.model_selection import train_test_split, cross_val_predict
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
import matplotlib.pyplot as plt
from sklearn.neighbors import KNeighborsRegressor
from mpl_toolkits.mplot3d import Axes3D


def generate_row_data(S=20, H=20, P=10, occ=0.8):
    yard = [[] for _ in range(S)]
    while sum([len(stack) for stack in yard]) < S * H * occ:
        stack_index = random.sample(range(S), 1)[0]
        if len(yard[stack_index]) == H: continue

        plate = random.sample(range(1, P + 1), 1)[0]
        yard[stack_index].append(plate)
    relocations = extended_LA_for_heu(yard, S, H)
    return yard, relocations


def data_save(yard, relocations, index=0, S=20, H=20):
    #print(yard)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'storage'
    for s in range(S):
        for h in range(H):
            if len(yard[s]) > h:
                sheet.cell(row=s + 1, column=h + 1).value = yard[s][h]

    workbook.create_sheet('relocations')
    sheet = workbook['relocations']
    for s in range(S):
        for h in range(H):
            if len(yard[s]) > h:
                sheet.cell(row=s + 1, column=h + 1).value = relocations[s][h]

    workbook.save(f'heu_instance/instance_{index}.xlsx')


def feature_extraction(url, feature_frame, S=20, H=20, P=10):
    # retrieve data
    workbook = openpyxl.load_workbook(url)
    storage_sheet = workbook['storage']
    yard = []
    for row in storage_sheet.rows:
        stack = []
        for cell in row:
            if cell.value is not None and cell.value != '':
                stack.append(int(cell.value))
        yard.append(stack)

    relocation_sheet = workbook['relocations']
    relocations = []
    for row in relocation_sheet.rows:
        stack = []
        for cell in row:
            if cell.value is not None and cell.value != '':
                stack.append(int(cell.value))
        relocations.append(stack)
    # how many plates in each group
    plate_num_in_group = [0 for _ in range(P)]
    for stack in yard:
        for plate in stack:
            plate_num_in_group[plate - 1] += 1

    for s in range(S):
        stack_group, first_nonBlocking_plate_slot = P, 0
        for h in range(len(yard[s])):
            # add the relocation number
            feature_list = [relocations[s][h]]
            # check if it's a blocking plate
            if yard[s][h] > stack_group:
                feature_list.append(1)
            else:
                feature_list.append(0)
                stack_group = yard[s][h]
                first_nonBlocking_plate_slot = h
            # add the group
            feature_list.append(float(yard[s][h] / P))
            # add the slot
            feature_list.append(float((h + 1) / H))
            # add the upper plate number
            feature_list.append(float((len(yard[s]) - h - 1) / H))
            # add the group gap to the first bottom non-blocking plate
            feature_list.append(float((yard[s][h] - stack_group) / P))
            # add the slot gap to the first bottom non-blocking plate
            feature_list.append(float((h - first_nonBlocking_plate_slot) / H))
            # add plates_num_in_between
            if stack_group + 1 >= yard[s][h]:
                feature_list.append(0)
            else:
                a = plate_num_in_group[stack_group:yard[s][h] - 1]
                feature_list.append(sum(a)/S)
            # add plates num in stack group
            feature_list.append(plate_num_in_group[stack_group - 1]/S)
            # add upper smaller block plate
            if h == len(yard[s]) - 1:
                feature_list.append(0)
                feature_list.append(0)
            else:
                temp_stack_group, smaller_grop_num, large_group_num = stack_group, 0, 0
                for plate in yard[s][h+1:len(yard[s])]:
                    if plate <= temp_stack_group: temp_stack_group = plate
                    elif plate < yard[s][h]: smaller_grop_num += 1
                    elif plate > yard[s][h]: large_group_num += 1
                feature_list.append(smaller_grop_num/S)
                feature_list.append(large_group_num/S)
            # add plate_num_in_between_in_lower_slot
            num = 0
            for plate in yard[s][:h+1]:
                if plate < yard[s][h]:
                    num += 1
            feature_list.append(num)
            # add plate_num_in_larger_group
            if yard[s][h] == P:
                feature_list.append(0)
            else:
                feature_list.append(sum(plate_num_in_group[yard[s][h]:])/S)
            # add the averaged stack group
            feature_list.append(sum(yard[s])/S)
            # add feature_list to dataframe
            feature_frame.loc[len(feature_frame)] = feature_list
    return feature_frame


def analysis():
    data = pd.read_csv('heu_instance/features.csv', index_col=0)
    print('There are',len(data), 'data\n')
    if input('Want to know Q1: non_blocking plate be relocated?') == '1':
        df = data[(data['relocation_time'] > 0) & (data['is_block_plate'] == 0)]
        if len(df) == 0:
            print('no')
        else:
            print(f'yes, {len(df)} non_blocking plates are relocated')
    data = data.drop(data[data['is_block_plate'] == 0].index)
    print('\nexclude non-blocking plates, remain', len(data), 'data\n')

    if input('Want to know Q2: the pearson and R ?') == '1':
        corr = data.corr(method='spearman')
        corr_to_relocation = corr['relocation_time'].drop(['relocation_time', 'is_block_plate'])
        print(corr_to_relocation)
        # fig = plt.figure()
        # ax = plt.axes(projection='3d')
        # ax.scatter3D(data['groupGap_to_unblock_plate'], data['group'], data['relocation_time'])
        # plt.show()

    if input('Want to know Q3: Linear regression ?') == '1':
        x = data[['group', 'groupGap_to_unblock_plate', 'plate_num_in_between', 'plate_num_in_larger_group']]
        y = data[['relocation_time']]
        lin_reg = LinearRegression()
        x_train, x_test, y_train, y_test = train_test_split(x, y, random_state=42)
        lin_reg.fit(x_train, y_train)
        print('linear coefficient:', lin_reg.coef_)
        print('linear intercept:', lin_reg.intercept_)
        print("Linear Model MSE:", mean_squared_error(y_test, lin_reg.predict(x_test)))


if __name__ == '__main__':
    # row data generation, use only once
    # for i in range(1000):
    #     yard, relocations = generate_row_data()
    #     data_save(yard, relocations, index=i)

    # feature_dataframe = pd.DataFrame(columns=['relocation_time',
    #                                           'is_block_plate',
    #                                           'group',
    #                                           'slot',
    #                                           'upper_plate_num',
    #                                           'groupGap_to_unblock_plate',
    #                                           'slotGap_to_unblock_plate',
    #                                           'plate_num_in_between',
    #                                           'plate_num_in_group_of_unblock_plate',
    #                                           'upper_smaller_block_plate',
    #                                           'upper_large_block_plate',
    #                                           'plate_num_in_between_in_lower_slot',
    #                                           'plate_num_in_larger_group',
    #                                           'the_averaged_stack_group'])
    # for file in os.listdir('heu_instance'):
    #     if file.endswith('.xlsx'):
    #         feature_dataframe = feature_extraction('heu_instance/'+file, feature_dataframe)
    #         print(file)
    #         # break
    # feature_dataframe.to_csv('heu_instance/features.csv')

    analysis()
