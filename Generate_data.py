'''
Generate instances given M, mean and standard deviations of the pre-process time
'''
import openpyxl
import random
import numpy as np
from itertools import product
from Utiles import parameter_setting

class instances_generator:
    def __init__(self, M, mean, std):
        self.M = M
        self.mean = mean
        self.std = std
        settings = parameter_setting()
        self.S = settings['S']
        self.H = settings['H']
        self.occ = settings['occ']
        self.plate_probability = settings['probability']
        self.P = settings['P']

    # return one group of plate according to the distribution of the group
    def _plate_generate(self):
        seed = random.random()
        for i in range(len(self.plate_probability)):
            if seed <= sum(self.plate_probability[:i+1]):
                return i + 1
        return len(self.plate_probability)

    def generate_configuration(self):
        configuration = [[] for _ in range(self.S)]
        for i in range(int(self.S * self.H * self.occ)):
            plate = self._plate_generate()
            s = random.randint(1, self.S)
            while len(configuration[s-1]) == self.H:
                s = random.randint(1, self.S)
            configuration[s-1].append(plate)
        return configuration

    def generate_inbound(self):
        inbound_plate = []
        preprocess = []
        while len(preprocess) < self.M:
            inbound_plate.append(self._plate_generate())
            preprocess.append(max(10, round(np.random.normal(self.mean, self.std, 1)[0], 2)))
        return inbound_plate, preprocess

    def generate_instance_workbook(self, inbound_plate, preprocess):
        workbook = openpyxl.Workbook()

        # sheet basic
        sheet1 = workbook.active
        sheet1.title = "Basic"
        sheet1.cell(row=1, column=1).value = 'N'
        sheet1.cell(row=2, column=1).value = 'S'
        sheet1.cell(row=3, column=1).value = 'H'
        sheet1.cell(row=4, column=1).value = 'P'
        sheet1.cell(row=5, column=1).value = 'N_i'
        sheet1.cell(row=6, column=1).value = 'M'
        sheet1.cell(row=7, column=1).value = 'R'
        sheet1.cell(row=8, column=1).value = 'W'
        sheet1.cell(row=9, column=1).value = 'eta'
        sheet1.cell(row=10, column=1).value = 'time_for_row_empty'
        sheet1.cell(row=11, column=1).value = 'time_for_column_empty'
        sheet1.cell(row=12, column=1).value = 'time_for_row_loaded'
        sheet1.cell(row=13, column=1).value = 'time_for_column_loaded'

        sheet1.cell(row=1, column=2).value = int(self.S * self.H * self.occ) + self.M
        sheet1.cell(row=2, column=2).value = self.S
        sheet1.cell(row=3, column=2).value = self.H
        sheet1.cell(row=4, column=2).value = self.P
        sheet1.cell(row=5, column=2).value = int(self.S * self.H * self.occ)
        sheet1.cell(row=6, column=2).value = self.M
        sheet1.cell(row=7, column=2).value = 4
        sheet1.cell(row=8, column=2).value = 5
        sheet1.cell(row=9, column=2).value = 0.5
        sheet1.cell(row=10, column=2).value = 0.25
        sheet1.cell(row=11, column=2).value = 0.25
        sheet1.cell(row=12, column=2).value = 0.125
        sheet1.cell(row=13, column=2).value = 0.125

        # sheet inbound
        sheet2 = workbook.create_sheet('Inbound')
        sheet2.cell(row=1, column=1).value = 'id'
        sheet2.cell(row=1, column=2).value = 'group'
        sheet2.cell(row=1, column=3).value = 'preprocess_time'
        for i in range(self.M):
            sheet2.cell(row = i+2, column=1).value = i + 1
            sheet2.cell(row = i+2, column=2).value = inbound_plate[i]
            sheet2.cell(row = i+2, column=3).value = preprocess[i]

        # sheet configuration
        sheet3 = workbook.create_sheet('Configuration')
        configuration = self.generate_configuration()
        for i, stack in enumerate(configuration):
            for j, plate in enumerate(stack):
                sheet3.cell(row=i+1, column=j+1).value = plate
        return workbook

if __name__ == '__main__':
    settings = parameter_setting()
    M_list, mean_list, std_list = settings['M_list'], settings['mean_list'], settings['std_list']
    for M, mean, std in product(M_list, mean_list, std_list):
        print(M, mean, std)
        generator = instances_generator(M, mean, std)
        inbound_plate, preprocess = generator.generate_inbound()
        for ins_num in range(20):
            instance_workbook = generator.generate_instance_workbook(inbound_plate, preprocess)
            instance_workbook.save(f'instance/M{M}mean{mean}_{ins_num}.xlsx')
    # M = 30
    # mean = 40
    # std = 4
    # ins = 1
    # generator = instances_generator(M, mean, std)
    # inbound_plate, preprocess = generator.generate_inbound()
    # instance_workbook = generator.generate_instance_workbook(inbound_plate, preprocess)
    # instance_workbook.save(f'instance/M{M}mean{mean}std{std}_{ins}.xlsx')