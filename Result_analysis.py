import copy
import math
import os
import openpyxl
from Utiles import load_data
from Environment import Environment
def get_results(met):
    if met == 'GA':
        path = 'result/ga'
    elif met == 'ACO':
        path = 'result/aco'
    elif met == 'ACO_J':
        path = 'result/aco_jovanovic'

    result_files = os.listdir(path)
    res_time = {}
    res_obj = {} #最终出库翻板几次，即目标值。
    res_rel = {} #入库作业过程翻板了几次。
    res_bp_num = {}
    for file_name in result_files:
        print(f'------------Analyzing {met} result:{file_name}------------')
        key = (met, int(file_name[1:3]), int(file_name[7:9])) # method, m
        if key not in res_obj.keys():
            res_obj[key] = []
            res_rel[key] = []
            res_bp_num[key] = []
            res_time[key] = []

        wb = openpyxl.load_workbook(path+'/'+file_name)
        ds = wb.active
        obj1 = int(ds.cell(row=3, column=2).value)
        rel1 = int(ds.cell(row=3, column=3).value)
        bpnum1 = int(ds.cell(row=3, column=4).value)
        for row in range(4, ds.max_row+1):
            obj1 = min(obj1, int(ds.cell(row=row, column=2).value))
            rel1 = max(rel1, int(ds.cell(row=row, column=3).value))
            bpnum1 = min(bpnum1, int(ds.cell(row=row, column=4).value))
        res_time[key].append(round(float(ds.cell(row=1, column=2).value), 2))
        res_obj[key].append(obj1)
        res_rel[key].append(rel1)
        res_bp_num[key].append(bpnum1)
    for key in res_time.keys():
        print(f'\nmethod:{key[0]}, M:{key[1]}, p:{key[2]}')
        print(f'time:{sum(res_time[key]) / len(res_time[key])}')
        print(f'obj:{sum(res_obj[key]) / len(res_obj[key])}')
        print(f'rel:{sum(res_rel[key]) / len(res_rel[key])}')
        print(f'bp_num:{sum(res_bp_num[key]) / len(res_bp_num[key])}')

def get_iteration_result(met):
    if met == 'GA':
        path = 'result/ga'
    elif met == 'ACO':
        path = 'result/aco'
    elif met == 'ACO_J':
        path = 'result/aco_jovanovic'

    result_files = os.listdir(path)
    res_0 = {}
    res_10 = {} #最终出库翻板几次，即目标值。
    res_50 = {} #入库作业过程翻板了几次。
    res_100 = {}
    for file_name in result_files:
        print(f'------------Analyzing {met} result:{file_name}------------')
        key = (met, int(file_name[1:3]), int(file_name[7:9])) # method, m
        if key not in res_0.keys():
            res_0[key] = []
            res_10[key] = []
            res_50[key] = []
            res_100[key] = []

        wb = openpyxl.load_workbook(path+'/'+file_name)
        ds = wb.active

        obj = math.inf
        for row in range(3, ds.max_row+1):
            obj = min(obj, int(ds.cell(row=row, column=2).value))
            if row == 3:
                res_0[key].append(obj)
            if row == 62:
                res_10[key].append(obj)
            if row == 302:
                res_50[key].append(obj)
            if row == 602:
                res_100[key].append(obj)
    for key in res_0.keys():
        init_res = sum(res_0[key])/len(res_0[key])
        per10_res = sum(res_10[key])/len(res_10[key])
        per50_res = sum(res_50[key]) / len(res_50[key])
        per100_res = sum(res_100[key]) / len(res_100[key])
        print(f'\nmethod:{key[0]}, M:{key[1]}, p:{key[2]}')
        print(f'res 0:{init_res}')
        print(f'res 10%:{per10_res}, rate{round(per10_res/init_res*100,2)}%')
        print(f'res 50%:{per50_res}, rate{round(per50_res / init_res * 100, 2)}%')
        print(f'res 100%:{per100_res}, rate{round(per100_res / init_res * 100, 2)}%')

def curve_convergence(met):
    if met == 'GA':
        path = 'result/ga'
    elif met == 'ACO':
        path = 'result/aco'
    elif met == 'ACO_J':
        path = 'result/aco_jovanovic'

    result_files = os.listdir(path)
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing {met} result:{file_name}------------')
        key = (met, int(file_name[1:3]))  # method, m

        wb = openpyxl.load_workbook(path + '/' + file_name)
        ds = wb.active

        if key not in res.keys():
            res[key] = []

            obj = math.inf
            for row in range(3, ds.max_row + 1):
                obj = min(obj, int(ds.cell(row=row, column=2).value))
                res[key].append(obj)
        else:
            obj = math.inf
            for row in range(3, ds.max_row + 1):
                obj = min(obj, int(ds.cell(row=row, column=2).value))
                res[key][row - 3] += obj

    for key in res.keys():
        print(f'\nmethod:{key[0]}, M:{key[1]}')
        for iter_val in res[key]:
            print(f'{iter_val/100}')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key[1]
        for ix1, val in enumerate(res[key]):
            sheet.cell(row=ix1+2, column=ix+1).value = float(val/100)
    workbook.save(f'result/curve_{met}.xlsx')

def get_rel_baseLine():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[1:3])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        env = Environment(
            S=S,
            H=H,
            M=M,
            P=P,
            inbound_plate=inbound_plates,
            preprocess_time=preprocess_time,
            configuration=config,
            operate_time=operate_time
        )

        res[key] += env.relocation_baseLine



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/rel_baseLine_M.xlsx')

def get_total_preprocess_time():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[1:3])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        res[key] += sum(preprocess_time)



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/total_preprocess_time_M.xlsx')

def get_initial_blocking_number():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[1:3])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        b_num = M
        for stack in config:
            min_g = math.inf
            for plate in stack:
                if plate <= min_g:
                    min_g = plate
                else:
                    b_num += 1
        res[key] += b_num



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/initial_bpnum_M.xlsx')

def get_initial_blocking_number_P():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[7:9])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        b_num = M
        for stack in config:
            min_g = math.inf
            for plate in stack:
                if plate <= min_g:
                    min_g = plate
                else:
                    b_num += 1
        res[key] += b_num



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/initial_bpnum_P.xlsx')

def get_total_preprocess_time_P():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[7:9])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        res[key] += sum(preprocess_time)



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/total_preprocess_time_P.xlsx')

def get_rel_baseLine_P():
    result_files = os.listdir('instance/all')
    res = {}

    for file_name in result_files:
        print(f'------------Analyzing result:{file_name}------------')
        key = int(file_name[7:9])  # m

        if key not in res.keys():
            res[key] = 0

        # M, mean, std = int(file_name[1:3]), int(file_name[7:9]), int(file_name[12])
        S, H, M, N_i, P, inbound_plates, preprocess_time, config, operate_time, min_operate_time = load_data(
            f'instance/all/{file_name}')
        env = Environment(
            S=S,
            H=H,
            M=M,
            P=P,
            inbound_plate=inbound_plates,
            preprocess_time=preprocess_time,
            configuration=config,
            operate_time=operate_time
        )

        res[key] += env.relocation_baseLine



    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for ix, key in enumerate(res.keys()):
        sheet.cell(row=1, column=ix+1).value = key
        sheet.cell(row=2, column=ix+1).value = float(res[key]/100)
    workbook.save(f'result/rel_baseLine_P.xlsx')

def get_results_P(met):
    if met == 'GA':
        path = 'result/ga'
    elif met == 'ACO':
        path = 'result/aco'
    elif met == 'ACO_J':
        path = 'result/aco_jovanovic'

    result_files = os.listdir(path)
    res_time = {}
    res_obj = {} #最终出库翻板几次，即目标值。
    res_rel = {} #入库作业过程翻板了几次。
    res_bp_num = {}
    for file_name in result_files:
        print(f'------------Analyzing {met} result:{file_name}------------')
        key = int(file_name[7:9]) # method, m
        if key not in res_obj.keys():
            res_obj[key] = []
            res_rel[key] = []
            res_bp_num[key] = []
            res_time[key] = []

        wb = openpyxl.load_workbook(path+'/'+file_name)
        ds = wb.active
        obj1 = int(ds.cell(row=3, column=2).value)
        rel1 = int(ds.cell(row=3, column=3).value)
        bpnum1 = int(ds.cell(row=3, column=4).value)
        for row in range(4, ds.max_row+1):
            obj1 = min(obj1, int(ds.cell(row=row, column=2).value))
            rel1 = max(rel1, int(ds.cell(row=row, column=3).value))
            bpnum1 = min(bpnum1, int(ds.cell(row=row, column=4).value))
        res_time[key].append(round(float(ds.cell(row=1, column=2).value), 2))
        res_obj[key].append(obj1)
        res_rel[key].append(rel1)
        res_bp_num[key].append(bpnum1)
    for key in res_time.keys():
        print(f'\np:{key}')
        print(f'time:{sum(res_time[key]) / len(res_time[key])}')
        print(f'obj:{sum(res_obj[key]) / len(res_obj[key])}')
        print(f'rel:{sum(res_rel[key]) / len(res_rel[key])}')
        print(f'bp_num:{sum(res_bp_num[key]) / len(res_bp_num[key])}')
if __name__ == '__main__':
    met = 'ACO'
    #curve_convergence(met)

    get_results_P(met)